#!/usr/bin/env python3

import copy
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
import sys


def ci_value(lower, upper):
    return '{:,} - {:,}'.format(int(lower), int(upper))


def gain_value(raw, pcnt, prec=0):
    return '{:,.{}f} ({:0.2f}%)'.format(raw, prec, pcnt)


def gain_ci_value(lwr, upr, lwr_pcnt, upr_pcnt, prec=0):
    if lwr < 0 and upr < 0:
        # Negative gain (i.e., a decrease), so reverse bounds order.
        lwr, upr = upr, lwr
        lwr_pcnt, upr_pcnt = upr_pcnt, lwr_pcnt

    return '{:,.{}f} - {:,.{}f} ({:0.2f}%, {:0.2f}%)'.format(
        lwr, prec, upr, prec, lwr_pcnt, upr_pcnt)


def tabulate_mm(ws, dfs):
    column_names = [
        'Scenario',
        'Age',
        'ACMR',
        'Pr(Death)',
        'Population',
        'Deaths',
        'Survivors',
        'Person Years',
        'LE',
        'YLD Rate',
        'HALYs',
        'HALE',
    ]
    for column_ix, column_name in enumerate(column_names):
        _ = ws.cell(column=column_ix + 1, row=1, value=column_name)

    df_acmr = dfs['reduce_acmr']
    df_chd = dfs['reduce_chd']

    sex = 'male'
    year_of_birth = 1959
    ages = [52, 53, 109, 110]

    df_acmr = df_acmr.loc[(df_acmr['sex'] == sex)
                          & (df_acmr['year_of_birth'] == year_of_birth)
                          & (df_acmr['age'].isin(ages))]
    df_chd = df_chd.loc[(df_chd['sex'] == sex)
                          & (df_chd['year_of_birth'] == year_of_birth)
                          & (df_chd['age'].isin(ages))]

    df_acmr = df_acmr.reset_index(drop=True)
    df_chd = df_chd.reset_index(drop=True)

    row = 2
    rows_per_scenario = df_acmr.shape[0]

    # Print life-table rows for the BAU scenario.
    for ix in range(rows_per_scenario):
        if ix == 0:
            ws.cell(column=1, row=row, value='BAU')
        ws.cell(column=2, row=row, value=df_acmr.loc[ix, 'age'])
        ws.cell(column=3, row=row, value=df_acmr.loc[ix, 'bau_acmr'])
        ws.cell(column=4, row=row, value=df_acmr.loc[ix, 'bau_pr_death'])
        ws.cell(column=5, row=row, value=df_acmr.loc[ix, 'bau_prev_population'])
        ws.cell(column=6, row=row, value=df_acmr.loc[ix, 'bau_deaths'])
        ws.cell(column=7, row=row, value=df_acmr.loc[ix, 'bau_population'])
        ws.cell(column=8, row=row, value=df_acmr.loc[ix, 'bau_person_years'])
        ws.cell(column=9, row=row, value=df_acmr.loc[ix, 'bau_LE'])
        ws.cell(column=10, row=row, value=df_acmr.loc[ix, 'bau_yld_rate'])
        ws.cell(column=11, row=row, value=df_acmr.loc[ix, 'bau_HALY'])
        ws.cell(column=12, row=row, value=df_acmr.loc[ix, 'bau_HALE'])
        row += 1

    # Print life-table rows for the reduced-ACMR scenario.
    for ix in range(rows_per_scenario):
        if ix == 0:
            ws.cell(column=1, row=row, value='Reduce ACMR by 5%')
        ws.cell(column=2, row=row, value=df_acmr.loc[ix, 'age'])
        ws.cell(column=3, row=row, value=df_acmr.loc[ix, 'acmr'])
        ws.cell(column=4, row=row, value=df_acmr.loc[ix, 'pr_death'])
        ws.cell(column=5, row=row, value=df_acmr.loc[ix, 'prev_population'])
        ws.cell(column=6, row=row, value=df_acmr.loc[ix, 'deaths'])
        ws.cell(column=7, row=row, value=df_acmr.loc[ix, 'population'])
        ws.cell(column=8, row=row, value=df_acmr.loc[ix, 'person_years'])
        ws.cell(column=9, row=row, value=df_acmr.loc[ix, 'LE'])
        ws.cell(column=10, row=row, value=df_acmr.loc[ix, 'yld_rate'])
        ws.cell(column=11, row=row, value=df_acmr.loc[ix, 'HALY'])
        ws.cell(column=12, row=row, value=df_acmr.loc[ix, 'HALE'])
        row += 1

    # Print life-table rows for the reduced-CHD scenario.
    for ix in range(rows_per_scenario):
        if ix == 0:
            ws.cell(column=1, row=row, value='Reduce CHD by 5%')
        ws.cell(column=2, row=row, value=df_chd.loc[ix, 'age'])
        ws.cell(column=3, row=row, value=df_chd.loc[ix, 'acmr'])
        ws.cell(column=4, row=row, value=df_chd.loc[ix, 'pr_death'])
        ws.cell(column=5, row=row, value=df_chd.loc[ix, 'prev_population'])
        ws.cell(column=6, row=row, value=df_chd.loc[ix, 'deaths'])
        ws.cell(column=7, row=row, value=df_chd.loc[ix, 'population'])
        ws.cell(column=8, row=row, value=df_chd.loc[ix, 'person_years'])
        ws.cell(column=9, row=row, value=df_chd.loc[ix, 'LE'])
        ws.cell(column=10, row=row, value=df_chd.loc[ix, 'yld_rate'])
        ws.cell(column=11, row=row, value=df_chd.loc[ix, 'HALY'])
        ws.cell(column=12, row=row, value=df_chd.loc[ix, 'HALE'])
        row += 1

    # Define how the quantities in each column should be displayed.
    num_formats = {
        'C': '0.0000', # ACMR
        'D': '0.0000', # P(death)
        'E': '#,##0',  # Initial population
        'F': '#,##0',  # Deaths
        'G': '#,##0',  # Survivors
        'H': '#,##0',  # Person-years
        'I': '0.00',   # Life expectancy
        'J': '0.0000', # YLD rate
        'K': '#,##0',  # Health-adjusted life-years
        'L': '0.00',   # Health-adjusted life expectancy
    }

    change_cols = ws['C:L']
    for col in change_cols:
        for cell in col[1:]:
            cell.number_format = num_formats[cell.column]


def tabulate(ws, start_row, delay, tob_prev, dfs):
    row = start_row

    ly_demographics = ['Total', 'Maori female', 'Non-Maori male']
    ly_popn = ['All', 'maori', 'non-maori']
    ly_sex = ['All', 'female', 'male']
    rate_demographics = ['Maori female', 'Maori female',
                         'Non-Maori male', 'Non-Maori male']
    rate_popn = ['maori', 'maori', 'non-maori', 'non-maori']
    rate_sex = ['female', 'female', 'male', 'male']
    rate_age = 62
    rate_years = [2041, 2061, 2041, 2061]

    interventions = ['erad', 'tax', 'tfg']

    # Extract the relevant subsets of each data table.
    df_ly = dfs['ly']
    df_ly = df_ly.loc[(df_ly['delay'] == delay)
                      & (df_ly['tob_prev'] == tob_prev)]
    df_haly = dfs['haly']
    df_haly = df_haly.loc[(df_haly['delay'] == delay)
                      & (df_haly['tob_prev'] == tob_prev)]
    df_acmr = dfs['acmr']
    df_acmr = df_acmr.loc[(df_acmr['delay'] == delay)
                      & (df_acmr['tob_prev'] == tob_prev)]
    df_yldr = dfs['yldr']
    df_yldr = df_yldr.loc[(df_yldr['delay'] == delay)
                      & (df_yldr['tob_prev'] == tob_prev)]

    # Write the LY gains obtained with each intervention.
    ws['A{}'.format(row)] = 'LY'
    for ix, label in enumerate(ly_demographics):
        ws.cell(column=2, row=row + ix, value=label)

        mask = (df_ly['popn'] == ly_popn[ix]) & (df_ly['sex'] == ly_sex[ix])

        # Write BAU quantites.
        bau_median = df_ly.loc[mask, 'bau_LY.median'].values[0]
        median = ws.cell(column=5, row=row + ix, value=bau_median)
        median.number_format = '#,##0'

        # Write the gains obtained from each intervention.
        column = 6
        for interv in interventions:
            df_int = df_ly.loc[mask & (df_ly['interv'] == interv)]

            # Median gain.
            int_median = df_int['LY_gain.median'].values[0]
            int_pcnt = df_int['LY_pcnt.median'].values[0]
            ws.cell(column=column, row=row + ix,
                    value=gain_value(int_median, int_pcnt))
            column += 1

            # Gain CI.
            lwr = df_int['LY_gain.lower'].values[0]
            upr = df_int['LY_gain.upper'].values[0]
            lwr_pcnt = df_int['LY_pcnt.lower'].values[0]
            upr_pcnt = df_int['LY_pcnt.upper'].values[0]
            ws.cell(column=column, row=row + ix,
                    value=gain_ci_value(lwr, upr, lwr_pcnt, upr_pcnt))
            column += 1

    row += len(ly_demographics)

    # Write the HALY gains obtained with each intervention.
    ws['A{}'.format(row)] = 'HALY'
    for ix, label in enumerate(ly_demographics):
        ws.cell(column=2, row=row + ix, value=label)

        mask = (df_haly['popn'] == ly_popn[ix]) & (df_haly['sex'] == ly_sex[ix])

        # Write BAU quantites.
        bau_median = df_haly.loc[mask, 'bau_HALY.median'].values[0]
        median = ws.cell(column=5, row=row + ix, value=bau_median)
        median.number_format = '#,##0'

        # Write the gains obtained from each intervention.
        column = 6
        for interv in interventions:
            df_int = df_haly.loc[mask & (df_haly['interv'] == interv)]

            # Median gain.
            int_median = df_int['HALY_gain.median'].values[0]
            int_pcnt = df_int['HALY_pcnt.median'].values[0]
            ws.cell(column=column, row=row + ix,
                    value=gain_value(int_median, int_pcnt))
            column += 1

            # Gain CI.
            lwr = df_int['HALY_gain.lower'].values[0]
            upr = df_int['HALY_gain.upper'].values[0]
            lwr_pcnt = df_int['HALY_pcnt.lower'].values[0]
            upr_pcnt = df_int['HALY_pcnt.upper'].values[0]
            ws.cell(column=column, row=row + ix,
                    value=gain_ci_value(lwr, upr, lwr_pcnt, upr_pcnt))
            column += 1

    row += len(ly_demographics)

    # Write the YLDR decreases obtained with each intervention.
    ws['A{}'.format(row)] = 'YLDR'
    for ix, label in enumerate(rate_demographics):
        ws.cell(column=2, row=row + ix, value=label)
        ws.cell(column=3, row=row + ix, value=rate_age)
        ws.cell(column=4, row=row + ix, value=rate_years[ix])

        mask = ((df_yldr['popn'] == rate_popn[ix])
                & (df_yldr['sex'] == rate_sex[ix])
                & (df_yldr['age'] == rate_age)
                & (df_yldr['year'] == rate_years[ix]))

        # Write BAU quantites.
        bau_median = df_yldr.loc[mask, 'bau_yld_rate.median'].values[0]
        median = ws.cell(column=5, row=row + ix, value=bau_median)
        median.number_format = '0.0000'

        # Write the gains obtained from each intervention.
        column = 6
        for interv in interventions:
            df_int = df_yldr.loc[mask & (df_yldr['interv'] == interv)]

            # Median gain.
            int_median = df_int['yld_rate_gain.median'].values[0]
            int_pcnt = df_int['yld_rate_pcnt.median'].values[0]
            ws.cell(column=column, row=row + ix,
                    value=gain_value(int_median, int_pcnt, prec=4))
            column += 1

            # Gain CI.
            lwr = df_int['yld_rate_gain.lower'].values[0]
            upr = df_int['yld_rate_gain.upper'].values[0]
            lwr_pcnt = df_int['yld_rate_pcnt.lower'].values[0]
            upr_pcnt = df_int['yld_rate_pcnt.upper'].values[0]
            ws.cell(column=column, row=row + ix,
                    value=gain_ci_value(lwr, upr, lwr_pcnt, upr_pcnt, prec=4))
            column += 1

    row += len(rate_demographics)

    # Write the ACMR decreases obtained with each intervention.
    ws['A{}'.format(row)] = 'ACMR'
    for ix, label in enumerate(rate_demographics):
        ws.cell(column=2, row=row + ix, value=label)
        ws.cell(column=3, row=row + ix, value=rate_age)
        ws.cell(column=4, row=row + ix, value=rate_years[ix])

        mask = ((df_acmr['popn'] == rate_popn[ix])
                & (df_acmr['sex'] == rate_sex[ix])
                & (df_acmr['age'] == rate_age)
                & (df_acmr['year'] == rate_years[ix]))

        # Write BAU quantites.
        bau_median = df_acmr.loc[mask, 'bau_acmr.median'].values[0]
        median = ws.cell(column=5, row=row + ix, value=bau_median)
        median.number_format = '#,##0'

        # Write the gains obtained from each intervention.
        column = 6
        for interv in interventions:
            df_int = df_acmr.loc[mask & (df_acmr['interv'] == interv)]

            # Median gain.
            int_median = df_int['acmr_gain.median'].values[0]
            int_pcnt = df_int['acmr_pcnt.median'].values[0]
            ws.cell(column=column, row=row + ix,
                    value=gain_value(int_median, int_pcnt))
            column += 1

            # Gain CI.
            lwr = df_int['acmr_gain.lower'].values[0]
            upr = df_int['acmr_gain.upper'].values[0]
            lwr_pcnt = df_int['acmr_pcnt.lower'].values[0]
            upr_pcnt = df_int['acmr_pcnt.upper'].values[0]
            ws.cell(column=column, row=row + ix,
                    value=gain_ci_value(lwr, upr, lwr_pcnt, upr_pcnt))
            column += 1

    row += len(rate_demographics)

    # Return the index of the next empty row.
    return row


def write_tables(output_file):
    # Read the results of the uncertainty analysis.
    dfs = {
        'acmr': pd.read_csv('uncertainty-ACMR.csv'),
        'haly': pd.read_csv('uncertainty-HALY.csv'),
        'ly': pd.read_csv('uncertainty-LY.csv'),
        'yldr': pd.read_csv('uncertainty-YLDR.csv'),
        'reduce_acmr': pd.read_csv('mslt_reduce_acmr_mm.csv'),
        'reduce_chd': pd.read_csv('mslt_reduce_chd_mm.csv'),
    }

    # Create an Excel workbook to contain the tabulated results.
    wb = openpyxl.Workbook()
    tbl2 = wb.active
    tbl2.title = 'Table 2'
    tbl3 = wb.create_sheet('Table 3')
    tbl4a = wb.create_sheet('Table 4a (constant prevalence)')
    tbl4b = wb.create_sheet('Table 4b (immediate recovery)')

    # Write column names for each table.
    column_names = [
        'Output',
	    'Demographic',
	    'Age',
	    'Calendar Year',
	    'BAU',
	    'Tobacco Eradication',
	    'Tobacco Eradication CIs',
	    'Tobacco Tax',
	    'Tobacco Tax CIs',
	    'Tobacco-Free Generation',
	    'Tobacco-Free Generation CIs',
    ]

    for column_ix, column_name in enumerate(column_names):
        _ = tbl2.cell(column=column_ix + 1, row=1, value=column_name)
        _ = tbl3.cell(column=column_ix + 1, row=1, value=column_name)
        _ = tbl4a.cell(column=column_ix + 1, row=1, value=column_name)
        _ = tbl4b.cell(column=column_ix + 1, row=1, value=column_name)

    # Tabulate the results.
    tabulate_mm(tbl2, dfs)
    first_empty_row = 2
    tabulate(tbl3, first_empty_row, 20, 'decreasing', dfs)
    tabulate(tbl4a, first_empty_row, 20, 'constant', dfs)
    tabulate(tbl4b, first_empty_row, 0, 'decreasing', dfs)

    # Bold the first row and first column of each table.
    font = copy.copy(tbl3.row_dimensions[1].font)
    font.bold = True
    tbl2.row_dimensions[1].font = font
    tbl2.column_dimensions['A'].font = font
    tbl3.row_dimensions[1].font = font
    tbl3.column_dimensions['A'].font = font
    tbl4a.row_dimensions[1].font = font
    tbl4a.column_dimensions['A'].font = font
    tbl4b.row_dimensions[1].font = font
    tbl4b.column_dimensions['A'].font = font

    # Save the Excel workbook.
    wb.save(output_file)


def main(args=None):
    write_tables('tables.xlsx')
    return 0


if __name__ == '__main__':
    sys.exit(main())
